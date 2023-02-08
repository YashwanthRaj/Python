import openpyxl as xl  # importing the required package to perform operations on excel using python
from openpyxl.chart import BarChart, Reference    # We are importing classes - BarChart, Reference from openpyxl


wb = xl.load_workbook('transactions.xlsx') # importing the excel workbook
sheet = wb['Sheet1'] # Sheet name

# Now we import cell location from the sheet, there are two ways:-
cell1 = sheet['a1'] # input coordinates of cell
cell2 = sheet.cell(1,1) # input row and column number
print(cell1.value) # This will return the value of that cell
print(cell2.value) # Same as above

print(sheet.max_row)  # Will return the maximum number of rows present in the sheet

# Loop to iterate through cells of 3rd column and decrease value by 10%
for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9  # Holds the correct value
    corrected_price_cell = sheet.cell(row,4) # Holds the cell for including the correct value
    corrected_price_cell.value = corrected_price

# To create the Bar chart
values = Reference(sheet,
                   min_row = 2,
                   max_row = 4,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('transactions2.xlsx')